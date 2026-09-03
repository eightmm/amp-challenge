from __future__ import annotations

import csv
import io
import json
import math
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .curation import canonicalize_sequence
from .data_registry import load_registry
from .measurements import (
    HC50_MENTION,
    MIC_MENTION,
    ConcentrationInterval,
    parse_hc50_entries,
    parse_mic_entries,
    threshold_label,
)
from .physchem import describe
from .splits import (
    assign_cluster_folds,
    cluster_lookup,
    fold_task_counts,
    single_linkage_global_edit_clusters,
)
from .utils import (
    atomic_write_json,
    atomic_write_text,
    canonical_json_bytes,
    sha256_bytes,
    sha256_file,
)

PARSER_VERSION = "dramp-xlsx-v1"
PARSER_COMPONENTS = (
    "curation.py",
    "measurements.py",
    "physchem.py",
    "splits.py",
    "training_data.py",
)

MIC_FIELDS = (
    "measurement_id",
    "sequence_id",
    "sequence",
    "n_terminal",
    "c_terminal",
    "source_name",
    "source_record_id",
    "source_row",
    "organism_name",
    "strain_name",
    "organism_raw",
    "taxonomy_confidence",
    "gram",
    "resistance_profile",
    "relation",
    "value_original_low",
    "value_original_high",
    "unit_original",
    "unit_normalized",
    "lower_um",
    "upper_um",
    "lower_log2_um",
    "upper_log2_um",
    "active_le_4um",
    "active_le_16um",
    "active_le_64um",
    "replicate_count",
    "pubmed_id",
    "cluster_id",
    "fold",
    "split_role",
)

HC50_FIELDS = (
    "measurement_id",
    "sequence_id",
    "sequence",
    "n_terminal",
    "c_terminal",
    "source_name",
    "source_record_id",
    "source_row",
    "cell_source",
    "endpoint",
    "relation",
    "value_original_low",
    "value_original_high",
    "unit_original",
    "unit_normalized",
    "lower_um",
    "upper_um",
    "lower_log2_um",
    "upper_log2_um",
    "safe_ge_64um",
    "safe_ge_128um",
    "replicate_count",
    "pubmed_id",
    "cluster_id",
    "fold",
    "split_role",
)

SPLIT_FIELDS = (
    "sequence_id",
    "sequence",
    "cluster_id",
    "cluster_representative",
    "cluster_size",
    "fold",
    "split_role",
    "mic_measurements",
    "hc50_measurements",
)

QUARANTINE_FIELDS = (
    "source_record_id",
    "source_row",
    "stage",
    "reason",
    "raw_excerpt",
)


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and math.isnan(value):
        return ""
    return unicodedata.normalize("NFKC", str(value)).strip()


def _sequence_id(sequence: str) -> str:
    return sha256_bytes(sequence.encode("ascii"))


def _stable_id(prefix: str, payload: dict[str, Any]) -> str:
    return f"{prefix}-{sha256_bytes(canonical_json_bytes(payload))[:20]}"


def _float(value: float | None) -> str:
    return "" if value is None else format(value, ".12g")


def _terminal_states(row: dict[str, Any]) -> tuple[str, str, list[str]]:
    reasons: list[str] = []
    n_raw = _text(row.get("N-terminal_Modification")).lower()
    c_raw = _text(row.get("C-terminal_Modification")).lower()
    other_raw = _text(row.get("Other_Modifications")).lower()
    if n_raw == "free":
        n_terminal = "free"
    elif n_raw == "acetylation":
        n_terminal = "acetylated"
    else:
        n_terminal = "unsupported"
        reasons.append("unsupported_n_terminal")
    if c_raw == "free":
        c_terminal = "free"
    elif c_raw in {"amidation", "α-amidation"}:
        c_terminal = "amidated"
    else:
        c_terminal = "unsupported"
        reasons.append("unsupported_c_terminal")
    if other_raw not in {"", "none", "free", "not found"}:
        reasons.append("unsupported_other_modification")
    return n_terminal, c_terminal, reasons


def _eligible_record(
    row: dict[str, Any], *, min_length: int, max_length: int
) -> tuple[str | None, str, str, list[str]]:
    reasons: list[str] = []
    try:
        sequence = canonicalize_sequence(_text(row.get("Sequence")))
    except ValueError:
        sequence = None
        reasons.append("noncanonical_sequence")
    if sequence is not None and not min_length <= len(sequence) <= max_length:
        reasons.append("length_out_of_range")
    shape = _text(row.get("Linear/Cyclic/Branched")).lower()
    if shape != "linear":
        reasons.append("not_linear")
    stereo = _text(row.get("Stereochemistry")).lower()
    if stereo != "l":
        reasons.append("not_all_l_stereochemistry")
    n_terminal, c_terminal, terminal_reasons = _terminal_states(row)
    reasons.extend(terminal_reasons)
    return sequence, n_terminal, c_terminal, reasons


def _interval_columns(interval: ConcentrationInterval) -> dict[str, str]:
    return {
        "relation": interval.relation,
        "value_original_low": _float(interval.value_original_low),
        "value_original_high": _float(interval.value_original_high),
        "unit_original": interval.unit_original,
        "unit_normalized": interval.unit_normalized,
        "lower_um": _float(interval.lower_um),
        "upper_um": _float(interval.upper_um),
        "lower_log2_um": _float(interval.lower_log2_um),
        "upper_log2_um": _float(interval.upper_log2_um),
    }


def _write_csv(path: Path, fields: tuple[str, ...], rows: list[dict[str, Any]]) -> None:
    buffer = io.StringIO(newline="")
    writer = csv.DictWriter(buffer, fieldnames=fields, lineterminator="\n", extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)
    atomic_write_text(path, buffer.getvalue())


def _add_quarantine(
    rows: list[dict[str, Any]],
    *,
    source_record_id: str,
    source_row: int,
    stage: str,
    reason: str,
    raw: str,
) -> None:
    excerpt = re.sub(r"\s+", " ", raw).strip()[:500]
    rows.append(
        {
            "source_record_id": source_record_id,
            "source_row": source_row,
            "stage": stage,
            "reason": reason,
            "raw_excerpt": excerpt,
        }
    )


def _deduplicate(rows: list[dict[str, Any]]) -> tuple[list[dict[str, Any]], int]:
    unique: dict[str, dict[str, Any]] = {}
    duplicates = 0
    for row in rows:
        identifier = row["measurement_id"]
        if identifier in unique:
            unique[identifier]["replicate_count"] = int(unique[identifier]["replicate_count"]) + 1
            duplicates += 1
        else:
            unique[identifier] = row
    return sorted(unique.values(), key=lambda row: row["measurement_id"]), duplicates


def _role(fold: int, config: dict[str, Any]) -> str:
    holdout = config["splits"]["holdout_folds"]
    if fold == int(holdout["test"]):
        return "test"
    if fold == int(holdout["calibration"]):
        return "calibration"
    return "train"


def prepare_training_data(
    *,
    project_root: Path,
    config_path: Path,
    registry_path: Path,
    output_dir: Path | None = None,
    raw_path: Path | None = None,
) -> dict[str, Any]:
    config = json.loads(config_path.read_text(encoding="utf-8"))
    if config.get("schema_version") != 1:
        raise ValueError(f"{config_path}: unsupported training-data config")
    registry = load_registry(registry_path)
    source_name = config["source"]
    source = registry["sources"][source_name]
    if source.get("access") != "automatic":
        raise PermissionError(f"{source_name}: source is not approved for automatic processing")
    source_path = raw_path or project_root / source["destination"]
    if not source_path.is_file():
        raise FileNotFoundError(
            f"missing {source_path}; run `uv run amp data fetch {source_name}` first"
        )
    raw_sha = sha256_file(source_path)
    if raw_sha != source["sha256"]:
        raise ValueError(f"{source_path}: SHA-256 {raw_sha} does not match the registry")
    destination = output_dir or project_root / config["output_dir"]
    destination.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(source_path, read_only=True, data_only=True)
    sheet_name = config["sheet"]
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"{source_path}: missing sheet {sheet_name!r}")
    sheet = workbook[sheet_name]
    row_values = sheet.iter_rows(values_only=True)
    try:
        headers = [_text(value) for value in next(row_values)]
    except StopIteration as error:
        raise ValueError(f"{source_path}: empty workbook") from error
    required = {
        "DRAMP_ID",
        "Sequence",
        "Activity",
        "Target_Organism",
        "Hemolytic_activity",
        "Linear/Cyclic/Branched",
        "N-terminal_Modification",
        "C-terminal_Modification",
        "Other_Modifications",
        "Stereochemistry",
        "Pubmed_ID",
    }
    missing = required - set(headers)
    if missing:
        raise ValueError(f"{source_path}: missing required columns {sorted(missing)}")

    stats: Counter[str] = Counter()
    quarantine: list[dict[str, Any]] = []
    raw_mic_rows: list[dict[str, Any]] = []
    raw_hc50_rows: list[dict[str, Any]] = []
    length_config = config["sequence"]
    max_concentration = float(config["measurements"]["max_concentration_um"])
    allowed_taxonomy = set(config["measurements"]["taxonomy_confidence"])

    for source_row, values in enumerate(row_values, start=2):
        stats["raw_rows"] += 1
        row = dict(zip(headers, values, strict=False))
        source_record_id = _text(row.get("DRAMP_ID")) or f"row-{source_row}"
        target_text = _text(row.get("Target_Organism"))
        hemolysis_text = _text(row.get("Hemolytic_activity"))
        mic_mentions = len(MIC_MENTION.findall(target_text))
        hc50_mentions = len(HC50_MENTION.findall(hemolysis_text))
        if mic_mentions:
            stats["rows_with_mic_mentions"] += 1
            stats["mic_mentions"] += mic_mentions
        if hc50_mentions:
            stats["rows_with_hc50_mentions"] += 1
            stats["hc50_mentions"] += hc50_mentions
        if not mic_mentions and not hc50_mentions:
            continue

        sequence, n_terminal, c_terminal, reasons = _eligible_record(
            row,
            min_length=int(length_config["min"]),
            max_length=int(length_config["max"]),
        )
        if reasons:
            stats["ineligible_labeled_rows"] += 1
            _add_quarantine(
                quarantine,
                source_record_id=source_record_id,
                source_row=source_row,
                stage="eligibility",
                reason=";".join(sorted(set(reasons))),
                raw=target_text or hemolysis_text,
            )
            continue
        assert sequence is not None
        stats["eligible_labeled_rows"] += 1
        sequence_id = _sequence_id(sequence)
        common = {
            "sequence_id": sequence_id,
            "sequence": sequence,
            "n_terminal": n_terminal,
            "c_terminal": c_terminal,
            "source_name": source_name,
            "source_record_id": source_record_id,
            "source_row": source_row,
            "replicate_count": 1,
            "pubmed_id": _text(row.get("Pubmed_ID")),
        }

        mic_entries, mic_errors = parse_mic_entries(
            target_text,
            sequence,
            n_terminal=n_terminal,
            c_terminal=c_terminal,
            activity=_text(row.get("Activity")),
        )
        stats["parsed_mic_mentions"] += len(mic_entries)
        unparsed_mic = max(0, mic_mentions - len(mic_entries) - len(mic_errors))
        if unparsed_mic:
            stats["unparsed_mic_mentions"] += unparsed_mic
            _add_quarantine(
                quarantine,
                source_record_id=source_record_id,
                source_row=source_row,
                stage="mic_parse",
                reason=f"unparsed_mentions={unparsed_mic}",
                raw=target_text,
            )
        for error in mic_errors:
            stats["mic_parse_errors"] += 1
            _add_quarantine(
                quarantine,
                source_record_id=source_record_id,
                source_row=source_row,
                stage="mic_parse",
                reason=error,
                raw=target_text,
            )
        for entry in mic_entries:
            organism = entry.organism
            if not organism.is_bacterial:
                stats["non_bacterial_mic"] += 1
                continue
            if organism.taxonomy_confidence not in allowed_taxonomy:
                stats["low_confidence_taxonomy_mic"] += 1
                _add_quarantine(
                    quarantine,
                    source_record_id=source_record_id,
                    source_row=source_row,
                    stage="taxonomy",
                    reason="low_confidence_bacterial_taxonomy",
                    raw=organism.raw,
                )
                continue
            interval = entry.interval
            finite_bounds = [value for value in (interval.lower_um, interval.upper_um) if value]
            if not finite_bounds or max(finite_bounds) > max_concentration:
                stats["mic_out_of_range"] += 1
                _add_quarantine(
                    quarantine,
                    source_record_id=source_record_id,
                    source_row=source_row,
                    stage="mic_qc",
                    reason="converted_concentration_out_of_range",
                    raw=organism.raw,
                )
                continue
            identity_payload = {
                "task": "mic",
                "sequence": sequence,
                "n_terminal": n_terminal,
                "c_terminal": c_terminal,
                "source_record_id": source_record_id,
                "organism_raw": organism.raw,
                "relation": interval.relation,
                "low": interval.value_original_low,
                "high": interval.value_original_high,
                "unit": interval.unit_normalized,
            }
            normalized = {
                **common,
                "measurement_id": _stable_id("mic", identity_payload),
                "organism_name": organism.name,
                "strain_name": organism.strain or "",
                "organism_raw": organism.raw,
                "taxonomy_confidence": organism.taxonomy_confidence,
                "gram": organism.gram,
                "resistance_profile": organism.resistance_profile or "",
                **_interval_columns(interval),
                "active_le_4um": threshold_label(interval, 4.0),
                "active_le_16um": threshold_label(interval, 16.0),
                "active_le_64um": threshold_label(interval, 64.0),
            }
            raw_mic_rows.append(normalized)

        hc50_entries, hc50_errors = parse_hc50_entries(
            hemolysis_text,
            sequence,
            n_terminal=n_terminal,
            c_terminal=c_terminal,
        )
        stats["parsed_hc50_mentions"] += len(hc50_entries)
        unparsed_hc50 = max(0, hc50_mentions - len(hc50_entries) - len(hc50_errors))
        if unparsed_hc50:
            stats["unparsed_hc50_mentions"] += unparsed_hc50
            _add_quarantine(
                quarantine,
                source_record_id=source_record_id,
                source_row=source_row,
                stage="hc50_parse",
                reason=f"unparsed_mentions={unparsed_hc50}",
                raw=hemolysis_text,
            )
        for error in hc50_errors:
            stats["hc50_parse_errors"] += 1
            _add_quarantine(
                quarantine,
                source_record_id=source_record_id,
                source_row=source_row,
                stage="hc50_parse",
                reason=error,
                raw=hemolysis_text,
            )
        for entry in hc50_entries:
            interval = entry.interval
            finite_bounds = [value for value in (interval.lower_um, interval.upper_um) if value]
            if not finite_bounds or max(finite_bounds) > max_concentration:
                stats["hc50_out_of_range"] += 1
                continue
            identity_payload = {
                "task": "hc50",
                "sequence": sequence,
                "n_terminal": n_terminal,
                "c_terminal": c_terminal,
                "source_record_id": source_record_id,
                "cell_source": entry.cell_source,
                "relation": interval.relation,
                "low": interval.value_original_low,
                "high": interval.value_original_high,
                "unit": interval.unit_normalized,
            }
            normalized = {
                **common,
                "measurement_id": _stable_id("hc50", identity_payload),
                "cell_source": entry.cell_source or "",
                "endpoint": "HC50",
                **_interval_columns(interval),
                "safe_ge_64um": threshold_label(interval, 64.0, higher_is_one=True),
                "safe_ge_128um": threshold_label(interval, 128.0, higher_is_one=True),
            }
            raw_hc50_rows.append(normalized)

    workbook.close()
    mic_rows, mic_duplicates = _deduplicate(raw_mic_rows)
    hc50_rows, hc50_duplicates = _deduplicate(raw_hc50_rows)
    stats["accepted_mic_measurements"] = len(mic_rows)
    stats["accepted_hc50_measurements"] = len(hc50_rows)
    stats["collapsed_mic_duplicates"] = mic_duplicates
    stats["collapsed_hc50_duplicates"] = hc50_duplicates
    sequences = sorted({row["sequence"] for row in mic_rows + hc50_rows})
    if not mic_rows:
        raise ValueError("no MIC measurements survived preparation")
    if not hc50_rows:
        raise ValueError("no HC50 measurements survived preparation")

    task_weights: dict[str, tuple[int, int]] = {}
    mic_counts = Counter(row["sequence"] for row in mic_rows)
    hc50_counts = Counter(row["sequence"] for row in hc50_rows)
    for sequence in sequences:
        task_weights[sequence] = (mic_counts[sequence], hc50_counts[sequence])
    cluster_config = config["clustering"]
    if cluster_config["method"] != "global_edit_single_linkage":
        raise ValueError(f"unsupported clustering method {cluster_config['method']!r}")
    clusters = single_linkage_global_edit_clusters(
        sequences,
        identity_threshold=float(cluster_config["identity_threshold"]),
        min_coverage=float(cluster_config["min_coverage"]),
    )
    cluster_ids = cluster_lookup(clusters)
    folds = assign_cluster_folds(
        clusters,
        task_weights=task_weights,
        folds=int(config["splits"]["folds"]),
        seed=int(config["splits"]["seed"]),
    )
    representatives = {
        member: cluster.representative for cluster in clusters for member in cluster.members
    }
    cluster_sizes = {
        member: len(cluster.members) for cluster in clusters for member in cluster.members
    }
    for row in mic_rows + hc50_rows:
        sequence = row["sequence"]
        row["cluster_id"] = cluster_ids[sequence]
        row["fold"] = folds[sequence]
        row["split_role"] = _role(folds[sequence], config)

    split_rows: list[dict[str, Any]] = []
    for sequence in sequences:
        fold = folds[sequence]
        split_rows.append(
            {
                "sequence_id": _sequence_id(sequence),
                "sequence": sequence,
                "cluster_id": cluster_ids[sequence],
                "cluster_representative": representatives[sequence],
                "cluster_size": cluster_sizes[sequence],
                "fold": fold,
                "split_role": _role(fold, config),
                "mic_measurements": mic_counts[sequence],
                "hc50_measurements": hc50_counts[sequence],
            }
        )

    feature_rows: list[dict[str, Any]] = []
    for sequence in sequences:
        feature_rows.append(
            {
                "sequence_id": _sequence_id(sequence),
                "sequence": sequence,
                **describe(sequence).as_dict(),
            }
        )
    feature_fields = tuple(feature_rows[0])
    quarantine.sort(
        key=lambda row: (
            int(row["source_row"]),
            str(row["stage"]),
            str(row["reason"]),
        )
    )

    output_rows = {
        "mic_measurements.csv": (MIC_FIELDS, mic_rows),
        "hc50_measurements.csv": (HC50_FIELDS, hc50_rows),
        "sequence_splits.csv": (SPLIT_FIELDS, split_rows),
        "sequence_features.csv": (feature_fields, feature_rows),
        "quarantine.csv": (QUARANTINE_FIELDS, quarantine),
    }
    for name, (fields, rows) in output_rows.items():
        _write_csv(destination / name, fields, rows)

    artifacts = {
        name: {"rows": len(rows), "sha256": sha256_file(destination / name)}
        for name, (_, rows) in output_rows.items()
    }
    fold_counts = fold_task_counts(folds, task_weights)
    manifest = {
        "schema_version": 1,
        "dataset_id": config["dataset_id"],
        "parser_version": PARSER_VERSION,
        "parser": {
            "version": PARSER_VERSION,
            "implementation_sha256": {
                name: sha256_file(Path(__file__).with_name(name)) for name in PARSER_COMPONENTS
            },
        },
        "source": {
            "name": source_name,
            "url": source["url"],
            "retrieved_at": source["retrieved_at"],
            "license": source["license"],
            "license_url": source["license_url"],
            "raw_path": str(source_path.relative_to(project_root))
            if source_path.is_relative_to(project_root)
            else str(source_path),
            "raw_sha256": raw_sha,
        },
        "config": {
            "path": str(config_path.relative_to(project_root)),
            "sha256": sha256_file(config_path),
        },
        "sequence_policy": config["sequence"],
        "clustering": {
            **cluster_config,
            "cluster_count": len(clusters),
            "largest_cluster": max(len(cluster.members) for cluster in clusters),
            "note": "Deterministic single-linkage global edit identity; not MMseqs2.",
        },
        "splits": {
            **config["splits"],
            "fold_counts": {str(key): value for key, value in fold_counts.items()},
        },
        "counts": dict(sorted(stats.items())),
        "artifacts": artifacts,
    }
    atomic_write_json(destination / "manifest.json", manifest)
    return manifest
